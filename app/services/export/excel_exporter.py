"""Export Excel (.xlsx) multi-onglets pour TAKA OS."""

import logging
from datetime import datetime
from typing import Optional
import uuid

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.ao_s2 import AO
from app.models.scoring import ScoringRun

logger = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


class ExcelExporter:
    """Exporteur Excel pour TAKA OS."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def export_full_report(
        self,
        tenant_id: uuid.UUID,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
    ) -> bytes:
        """Genere un rapport Excel complet multi-onglets."""
        wb = Workbook()
        wb.remove(wb.active)

        await self._sheet_aos(wb, tenant_id, date_from, date_to)
        await self._sheet_scoring(wb, tenant_id, date_from, date_to)
        await self._sheet_submissions(wb, tenant_id, date_from, date_to)
        await self._sheet_summary(wb, tenant_id, date_from, date_to)

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def _sheet_aos(self, wb, tenant_id, date_from, date_to):
        """Onglet Appels d'Offres."""
        ws = wb.create_sheet("Appels d'Offres")
        headers = ["ID", "Titre", "Reference", "CPV", "Statut", "Deadline", "Montant", "Devise", "Localisation", "Date creation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = BORDER

        conditions = [AO.tenant_id == tenant_id]
        if date_from: conditions.append(AO.created_at >= date_from)
        if date_to: conditions.append(AO.created_at <= date_to)

        stmt = select(AO).where(and_(*conditions)).order_by(AO.created_at.desc())
        result = await self.db.execute(stmt)
        aos = result.scalars().all()

        for row_idx, ao in enumerate(aos, 2):
            ws.cell(row=row_idx, column=1, value=str(ao.id)).border = BORDER
            ws.cell(row=row_idx, column=2, value=ao.title).border = BORDER
            ws.cell(row=row_idx, column=3, value=ao.external_id or "").border = BORDER
            ws.cell(row=row_idx, column=4, value="|".join(ao.cpv_codes) if ao.cpv_codes else "").border = BORDER
            ws.cell(row=row_idx, column=5, value=ao.status).border = BORDER
            ws.cell(row=row_idx, column=6, value=ao.deadline_date.strftime("%d/%m/%Y") if ao.deadline_date else "").border = BORDER
            ws.cell(row=row_idx, column=7, value=float(ao.estimated_amount) if ao.estimated_amount else "").border = BORDER
            ws.cell(row=row_idx, column=8, value=ao.currency).border = BORDER
            ws.cell(row=row_idx, column=9, value=ao.city or "").border = BORDER
            ws.cell(row=row_idx, column=10, value=ao.created_at.strftime("%d/%m/%Y %H:%M") if ao.created_at else "").border = BORDER

        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 20

    async def _sheet_scoring(self, wb, tenant_id, date_from, date_to):
        """Onglet Scoring."""
        ws = wb.create_sheet("Scoring")
        headers = ["AO ID", "Score Global", "Profil", "Dimensions", "Date"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER

        conditions = [ScoringRun.tenant_id == tenant_id]
        if date_from: conditions.append(ScoringRun.created_at >= date_from)
        if date_to: conditions.append(ScoringRun.created_at <= date_to)

        stmt = select(ScoringRun).where(and_(*conditions)).order_by(ScoringRun.created_at.desc()).limit(1000)
        result = await self.db.execute(stmt)
        scores = result.scalars().all()

        for row_idx, s in enumerate(scores, 2):
            ws.cell(row=row_idx, column=1, value=str(s.ao_id)).border = BORDER
            ws.cell(row=row_idx, column=2, value=float(s.score_global) if s.score_global else "").border = BORDER
            ws.cell(row=row_idx, column=3, value=s.profile or "").border = BORDER
            ws.cell(row=row_idx, column=4, value=str(s.dimensions) if s.dimensions else "").border = BORDER
            ws.cell(row=row_idx, column=5, value=s.created_at.strftime("%d/%m/%Y") if s.created_at else "").border = BORDER

    async def _sheet_submissions(self, wb, tenant_id, date_from, date_to):
        """Onglet Soumissions."""
        ws = wb.create_sheet("Soumissions")
        headers = ["ID", "Plateforme", "Statut", "Date"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER

        from app.models.submission import Submission
        conditions = [Submission.tenant_id == tenant_id]
        if date_from: conditions.append(Submission.created_at >= date_from)
        if date_to: conditions.append(Submission.created_at <= date_to)

        stmt = select(Submission).where(and_(*conditions)).order_by(Submission.created_at.desc()).limit(1000)
        result = await self.db.execute(stmt)
        subs = result.scalars().all()

        for row_idx, s in enumerate(subs, 2):
            ws.cell(row=row_idx, column=1, value=str(s.id)).border = BORDER
            ws.cell(row=row_idx, column=2, value=s.platform_reference or "").border = BORDER
            ws.cell(row=row_idx, column=3, value=s.status).border = BORDER
            ws.cell(row=row_idx, column=4, value=s.created_at.strftime("%d/%m/%Y") if s.created_at else "").border = BORDER

    async def _sheet_summary(self, wb, tenant_id, date_from, date_to):
        """Onglet Resume avec KPIs."""
        ws = wb.create_sheet("Resume")

        conditions = [AO.tenant_id == tenant_id]
        if date_from: conditions.append(AO.created_at >= date_from)
        if date_to: conditions.append(AO.created_at <= date_to)

        total_stmt = select(func.count(AO.id)).where(and_(*conditions))
        total_result = await self.db.execute(total_stmt)
        total = total_result.scalar()

        won_stmt = select(func.count(AO.id)).where(and_(*conditions, AO.status == "won"))
        won_result = await self.db.execute(won_stmt)
        won = won_result.scalar()

        value_stmt = select(func.sum(AO.estimated_amount)).where(
            and_(*conditions, AO.status == "won", AO.estimated_amount.isnot(None))
        )
        value_result = await self.db.execute(value_stmt)
        total_value = float(value_result.scalar() or 0)

        kpis = [
            ["Total AO", total],
            ["AO Gagnes", won],
            ["Taux de reussite", f"{(won/total*100):.1f}%" if total else "N/A"],
            ["Valeur totale gagnee", f"{total_value:,.0f} EUR"],
        ]

        for row_idx, (label, value) in enumerate(kpis, 1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
